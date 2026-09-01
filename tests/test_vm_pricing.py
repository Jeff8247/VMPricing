import unittest
from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

from vm_pricing import (
    AZURE_REGION,
    OS_LINUX,
    OS_RHEL,
    OS_WINDOWS,
    VmSize,
    build_row,
    cheapest_per_provider,
    fetch_aws_compute_price,
    parse_azure_sizes,
    parse_rba_usd_rate,
    select_azure_e10_price,
    select_azure_rhel_license_prices,
    select_azure_windows_prices,
    write_excel,
)


class PricingTests(unittest.TestCase):
    def test_aws_type_without_windows_meter_is_excluded(self):
        class EmptyPaginator:
            def paginate(self, **kwargs):
                return [{"PriceList": []}]

        class PricingClient:
            def get_paginator(self, name):
                self.assert_name = name
                return EmptyPaginator()

        self.assertIsNone(fetch_aws_compute_price(PricingClient(), "inf2.xlarge"))

    def test_rba_uses_latest_usd_observation(self):
        xml = """<rdf><item><targetCurrency>USD</targetCurrency><value>0.65</value><date>2026-08-28</date></item>
        <item><targetCurrency>USD</targetCurrency><value>0.66</value><date>2026-08-31</date></item>
        <item><targetCurrency>EUR</targetCurrency><value>0.60</value><date>2026-09-01</date></item></rdf>"""
        self.assertEqual(parse_rba_usd_rate(xml), (Decimal("0.66"), "2026-08-31"))

    def test_azure_sizes_require_exact_shape_and_availability(self):
        def sku(name, cpu, ram, architecture="x64", restrictions=None):
            return {
                "name": name,
                "resourceType": "virtualMachines",
                "locations": [AZURE_REGION],
                "capabilities": [
                    {"name": "vCPUs", "value": str(cpu)},
                    {"name": "MemoryGB", "value": str(ram)},
                    {"name": "CpuArchitectureType", "value": architecture},
                ],
                "restrictions": restrictions or [],
            }

        items = [
            sku("Standard_Good2", 2, 8),
            sku("Standard_Good4", 4, 16),
            sku("Standard_Wrong", 4, 8),
            sku("Standard_Arm", 2, 8, "Arm64"),
            sku("Standard_Blocked", 2, 8, restrictions=[{"type": "Location", "values": [AZURE_REGION]}]),
        ]
        self.assertEqual([item.name for item in parse_azure_sizes(items)], ["Standard_Good2", "Standard_Good4"])

    def test_azure_windows_selector_excludes_non_payg_rates(self):
        base = {
            "armSkuName": "Standard_D2s_v5",
            "productName": "Virtual Machines Dsv5 Series Windows",
            "meterName": "D2s v5",
            "skuName": "D2s v5",
            "type": "Consumption",
            "unitOfMeasure": "1 Hour",
            "retailPrice": 0.3,
            "effectiveStartDate": "2026-01-01T00:00:00Z",
        }
        spot = {**base, "meterName": "D2s v5 Spot", "retailPrice": 0.1}
        linux = {**base, "productName": "Virtual Machines Dsv5 Series", "retailPrice": 0.2}
        selected = select_azure_windows_prices([spot, linux, base])
        self.assertEqual(selected["Standard_D2s_v5"].value, Decimal("0.3"))

    def test_azure_ambiguous_compute_price_fails(self):
        one = {
            "armSkuName": "Standard_D2s_v5", "productName": "Virtual Machines Dsv5 Windows", "meterName": "D2s v5",
            "skuName": "D2s v5", "type": "Consumption", "unitOfMeasure": "1 Hour",
            "retailPrice": 0.3, "effectiveStartDate": "2026-01-01",
        }
        with self.assertRaisesRegex(Exception, "(?i)ambiguous"):
            select_azure_windows_prices([one, {**one, "retailPrice": 0.4}])

    def test_new_effective_price_supersedes_history(self):
        old = {
            "armSkuName": "Standard_D2s_v5", "productName": "Virtual Machines Dsv5 Windows", "meterName": "D2s v5",
            "skuName": "D2s v5", "type": "Consumption", "unitOfMeasure": "1 Hour",
            "retailPrice": 0.3, "effectiveStartDate": "2025-01-01",
        }
        selected = select_azure_windows_prices([old, {**old, "retailPrice": 0.4, "effectiveStartDate": "2026-01-01"}])
        self.assertEqual(selected["Standard_D2s_v5"].value, Decimal("0.4"))

    def test_expired_meter_is_ignored(self):
        expired = {
            "armSkuName": "Standard_D2s_v5", "productName": "Virtual Machines Dsv5 Windows", "meterName": "D2s v5",
            "skuName": "D2s v5", "type": "Consumption", "unitOfMeasure": "1 Hour",
            "retailPrice": 0.3, "effectiveStartDate": "2026-01-01", "effectiveEndDate": "2026-08-31",
        }
        current = {**expired, "retailPrice": 0.4, "effectiveEndDate": None}
        selected = select_azure_windows_prices([expired, current])
        self.assertEqual(selected["Standard_D2s_v5"].value, Decimal("0.4"))

    def test_e10_ignores_operations_meter(self):
        items = [
            {"meterName": "E10 LRS Disk Operations", "type": "Consumption", "unitOfMeasure": "10K", "retailPrice": 0.1},
            {"meterName": "E10 LRS Disk", "type": "Consumption", "unitOfMeasure": "1/Month", "retailPrice": 9.5,
             "effectiveStartDate": "2026-01-01"},
        ]
        self.assertEqual(select_azure_e10_price(items).value, Decimal("9.5"))

    def test_azure_rhel_license_is_selected_by_vcpu(self):
        items = []
        for vcpu, price in ((2, 0.04), (4, 0.08)):
            items.append({
                "meterName": f"{vcpu} vCPU VM License", "type": "Consumption", "unitOfMeasure": "1 Hour",
                "retailPrice": price, "effectiveStartDate": "2026-01-01",
            })
        items.append({
            "meterName": "2 vCPU VM BYOS License", "type": "Consumption", "unitOfMeasure": "1 Hour",
            "retailPrice": 0, "effectiveStartDate": "2026-01-01",
        })
        selected = select_azure_rhel_license_prices(items)
        self.assertEqual(selected[2].value, Decimal("0.04"))
        self.assertEqual(selected[4].value, Decimal("0.08"))

    def test_totals_include_amortised_disk(self):
        row = build_row("Azure", AZURE_REGION, VmSize("x", 2, Decimal("8")), Decimal("1"), "disk",
                        Decimal("73"), Decimal("730"), "2026-01-01")
        self.assertEqual(row.total_monthly_aud, Decimal("803"))
        self.assertEqual(row.total_hourly_aud, Decimal("1.1"))

    def test_excel_has_provider_sheets_and_numeric_prices(self):
        aws = build_row("AWS", "ap-southeast-2", VmSize("m.test", 2, Decimal("8")), Decimal("1.2"),
                        "128 GiB gp3", Decimal("10"), Decimal("730"), "2026-01-01", "0.7", "2026-01-01")
        azure = build_row("Azure", AZURE_REGION, VmSize("Standard_Test", 4, Decimal("16")), Decimal("1.1"),
                          "128 GiB E10", Decimal("12"), Decimal("730"), "2026-01-01")
        with TemporaryDirectory() as directory:
            path = Path(directory) / "prices.xlsx"
            write_excel([aws, azure], path)
            workbook = load_workbook(path)
            self.assertEqual(workbook.sheetnames, ["AWS", "Azure"])
            self.assertIn("one separate 128 GiB gp3", workbook["AWS"]["A1"].value)
            self.assertIn("one separate 128 GiB Standard SSD LRS", workbook["Azure"]["A1"].value)
            self.assertEqual(workbook["AWS"]["C4"].value, "2 vCPU / 8 GiB RAM")
            self.assertEqual(workbook["AWS"]["E4"].value, "m.test")
            self.assertIsInstance(workbook["AWS"]["H4"].value, float)
            self.assertEqual(workbook["AWS"].freeze_panes, "A4")

    def test_top_limit_is_applied_per_provider(self):
        rows = []
        for provider in ("AWS", "Azure"):
            for number in range(12):
                rows.append(build_row(
                    provider, "region", VmSize(f"size-{number}", 2, Decimal("8")), Decimal(number),
                    "disk", Decimal("0"), Decimal("730"), "2026-01-01",
                ))
        selected = cheapest_per_provider(rows, 10)
        self.assertEqual(sum(row.provider == "AWS" for row in selected), 10)
        self.assertEqual(sum(row.provider == "Azure" for row in selected), 10)
        self.assertNotIn("size-11", {row.instance_type for row in selected})

    def test_top_limit_is_independent_for_each_os(self):
        rows = []
        for operating_system in (OS_WINDOWS, OS_LINUX, OS_RHEL):
            for number in range(12):
                rows.append(build_row(
                    "AWS", "region", VmSize(f"{operating_system}-{number}", 2, Decimal("8")), Decimal(number),
                    "disk", Decimal("0"), Decimal("730"), "2026-01-01", operating_system=operating_system,
                ))
        selected = cheapest_per_provider(rows, 10)
        self.assertEqual(sum(row.operating_system == OS_WINDOWS for row in selected), 10)
        self.assertEqual(sum(row.operating_system == OS_LINUX for row in selected), 10)
        self.assertEqual(sum(row.operating_system == OS_RHEL for row in selected), 10)

    def test_top_limit_is_independent_for_each_vm_shape(self):
        rows = []
        for vcpu, memory in ((2, Decimal("8")), (4, Decimal("16"))):
            for number in range(7):
                rows.append(build_row(
                    "AWS", "region", VmSize(f"{vcpu}vcpu-{number}", vcpu, memory), Decimal(number),
                    "disk", Decimal("0"), Decimal("730"), "2026-01-01",
                ))
        selected = cheapest_per_provider(rows, 5)
        self.assertEqual(sum(row.vcpu == 2 and row.memory_gib == 8 for row in selected), 5)
        self.assertEqual(sum(row.vcpu == 4 and row.memory_gib == 16 for row in selected), 5)
        self.assertEqual(len(selected), 10)


if __name__ == "__main__":
    unittest.main()
