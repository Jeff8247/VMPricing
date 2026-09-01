#!/usr/bin/env python3
"""Compare live Windows, standard Linux, and RHEL VM prices in Sydney."""

from __future__ import annotations

import argparse
import os
import sys
import xml.etree.ElementTree as ET
from dataclasses import asdict, dataclass
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable, Iterator, Sequence

import boto3
import requests
from azure.identity import DefaultAzureCredential
from botocore.config import Config
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


AWS_REGION = "ap-southeast-2"
AZURE_REGION = "australiaeast"
TARGET_SHAPES = {(2, Decimal("8")), (4, Decimal("16"))}
RBA_RSS_URL = "https://www.rba.gov.au/rss/rss-cb-exchange-rates.xml"
AZURE_RETAIL_URL = "https://prices.azure.com/api/retail/prices"
AZURE_SKUS_API_VERSION = "2025-04-01"
DEFAULT_HOURS = Decimal("730")
HTTP_TIMEOUT = 60
OS_WINDOWS = "Windows"
OS_LINUX = "Linux (Ubuntu/Debian)"
OS_RHEL = "Red Hat Enterprise Linux"
ALL_OPERATING_SYSTEMS = (OS_WINDOWS, OS_LINUX, OS_RHEL)


class PricingError(RuntimeError):
    """An input feed was missing, ambiguous, or unusable."""


@dataclass(frozen=True)
class VmSize:
    name: str
    vcpu: int
    memory_gib: Decimal


@dataclass(frozen=True)
class Price:
    value: Decimal
    effective_date: str
    effective_end_date: str = ""


@dataclass
class ResultRow:
    provider: str
    region: str
    instance_type: str
    vcpu: int
    memory_gib: Decimal
    compute_hourly_aud: Decimal
    disk: str
    disk_monthly_aud: Decimal
    total_hourly_aud: Decimal
    total_monthly_aud: Decimal
    currency: str
    pricing_effective_date: str
    fx_aud_usd: str
    fx_date: str
    status: str = "ok"
    operating_system: str = OS_WINDOWS
    vm_shape: str = ""


def decimal(value: Any, context: str) -> Decimal:
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise PricingError(f"Invalid numeric value for {context}: {value!r}") from exc


def http_session() -> requests.Session:
    session = requests.Session()
    adapter = requests.adapters.HTTPAdapter(max_retries=3)
    session.mount("https://", adapter)
    session.headers["User-Agent"] = "sydney-windows-vm-pricing/1.0"
    return session


def get_json(session: requests.Session, url: str, **kwargs: Any) -> dict[str, Any]:
    response = session.get(url, timeout=HTTP_TIMEOUT, **kwargs)
    response.raise_for_status()
    payload = response.json()
    if not isinstance(payload, dict):
        raise PricingError(f"Expected an object from {url}")
    return payload


def local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def parse_rba_usd_rate(xml_text: str) -> tuple[Decimal, str]:
    """Return foreign currency units (USD) per one AUD and the observation date."""
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError as exc:
        raise PricingError("RBA exchange-rate feed was not valid XML") from exc

    observations: list[tuple[str, Decimal]] = []
    for item in root.iter():
        if local_name(item.tag) != "item":
            continue
        values: dict[str, list[str]] = {}
        for node in item.iter():
            if node.text and node.text.strip():
                values.setdefault(local_name(node.tag), []).append(node.text.strip())
        currencies = values.get("targetCurrency", [])
        if "USD" not in currencies:
            continue
        raw_values = values.get("value", [])
        dates = values.get("date", [])
        if raw_values and dates:
            observations.append((dates[-1][:10], decimal(raw_values[-1], "RBA AUD/USD rate")))

    if not observations:
        raise PricingError("RBA feed did not contain an AUD/USD observation")
    rate_date, rate = max(observations, key=lambda item: item[0])
    if rate <= 0:
        raise PricingError("RBA AUD/USD rate must be positive")
    return rate, rate_date


def fetch_rba_usd_rate(session: requests.Session) -> tuple[Decimal, str]:
    response = session.get(RBA_RSS_URL, timeout=HTTP_TIMEOUT)
    response.raise_for_status()
    rate, rate_date = parse_rba_usd_rate(response.text)
    age_days = (date.today() - date.fromisoformat(rate_date)).days
    if age_days < 0 or age_days > 7:
        raise PricingError(f"Latest RBA AUD/USD observation is stale or future-dated: {rate_date}")
    return rate, rate_date


def select_current_price(prices: Sequence[Price], context: str) -> Price:
    """Select the newest effective price on or before today, rejecting conflicting peers."""
    today = date.today().isoformat()
    current = [
        price for price in prices
        if price.effective_date
        and price.effective_date <= today
        and (not price.effective_end_date or price.effective_end_date >= today)
    ]
    if not current:
        raise PricingError(f"No currently effective rate for {context}")
    newest_date = max(price.effective_date for price in current)
    newest = [price for price in current if price.effective_date == newest_date]
    values = {price.value for price in newest}
    if len(values) != 1:
        raise PricingError(f"Ambiguous currently effective rates for {context}: {sorted(values)}")
    return newest[0]


def aws_clients(profile: str | None) -> tuple[Any, Any]:
    aws_session = boto3.Session(profile_name=profile) if profile else boto3.Session()
    retry_config = Config(retries={"max_attempts": 5, "mode": "standard"})
    return (
        aws_session.client("ec2", region_name=AWS_REGION, config=retry_config),
        aws_session.client("pricing", region_name="us-east-1", config=retry_config),
    )


def discover_aws_sizes(ec2: Any) -> list[VmSize]:
    offered: set[str] = set()
    offering_pages = ec2.get_paginator("describe_instance_type_offerings").paginate(
        LocationType="region", Filters=[{"Name": "location", "Values": [AWS_REGION]}]
    )
    for page in offering_pages:
        offered.update(item["InstanceType"] for item in page.get("InstanceTypeOfferings", []))

    sizes: list[VmSize] = []
    pages = ec2.get_paginator("describe_instance_types").paginate(
        Filters=[
            {"Name": "supported-usage-class", "Values": ["on-demand"]},
            {"Name": "vcpu-info.default-vcpus", "Values": ["2", "4"]},
        ]
    )
    for page in pages:
        for item in page.get("InstanceTypes", []):
            name = item["InstanceType"]
            vcpu = int(item["VCpuInfo"]["DefaultVCpus"])
            memory = decimal(item["MemoryInfo"]["SizeInMiB"], f"{name} memory") / 1024
            architectures = set(item.get("ProcessorInfo", {}).get("SupportedArchitectures", []))
            if name in offered and "x86_64" in architectures and (vcpu, memory) in TARGET_SHAPES:
                sizes.append(VmSize(name, vcpu, memory))
    return sorted(sizes, key=lambda item: (item.vcpu, item.name))


def aws_products(pricing: Any, filters: Sequence[dict[str, str]]) -> Iterator[dict[str, Any]]:
    paginator = pricing.get_paginator("get_products")
    import json

    for page in paginator.paginate(ServiceCode="AmazonEC2", Filters=list(filters), FormatVersion="aws_v1"):
        for encoded in page.get("PriceList", []):
            yield json.loads(encoded)


def aws_ondemand_price(product: dict[str, Any], expected_unit: str) -> Price | None:
    found: list[Price] = []
    for term in product.get("terms", {}).get("OnDemand", {}).values():
        effective = str(term.get("effectiveDate", ""))[:10]
        for dimension in term.get("priceDimensions", {}).values():
            if dimension.get("unit") != expected_unit:
                continue
            usd = decimal(dimension.get("pricePerUnit", {}).get("USD"), "AWS USD price")
            if usd >= 0 and str(dimension.get("beginRange", "0")) == "0":
                found.append(Price(usd, effective))
    if not found:
        return None
    return select_current_price(found, "AWS on-demand product")


def fetch_aws_compute_price(pricing: Any, instance_type: str, operating_system: str = OS_WINDOWS) -> Price | None:
    aws_os = {OS_WINDOWS: "Windows", OS_LINUX: "Linux", OS_RHEL: "RHEL"}[operating_system]
    filters = [
        {"Type": "TERM_MATCH", "Field": "regionCode", "Value": AWS_REGION},
        {"Type": "TERM_MATCH", "Field": "productFamily", "Value": "Compute Instance"},
        {"Type": "TERM_MATCH", "Field": "instanceType", "Value": instance_type},
        {"Type": "TERM_MATCH", "Field": "operatingSystem", "Value": aws_os},
        {"Type": "TERM_MATCH", "Field": "tenancy", "Value": "Shared"},
        {"Type": "TERM_MATCH", "Field": "preInstalledSw", "Value": "NA"},
        {"Type": "TERM_MATCH", "Field": "capacitystatus", "Value": "Used"},
    ]
    operation = {OS_WINDOWS: "RunInstances:0002", OS_LINUX: "RunInstances", OS_RHEL: "RunInstances:0010"}
    filters.append({"Type": "TERM_MATCH", "Field": "operation", "Value": operation[operating_system]})
    prices = [price for product in aws_products(pricing, filters) if (price := aws_ondemand_price(product, "Hrs"))]
    if not prices:
        # Some x86-64 accelerator types (for example AWS Inferentia) match the
        # requested hardware shape but do not support Windows Server. The live
        # catalog's lack of a Windows meter is the authoritative exclusion.
        return None
    return select_current_price(prices, f"AWS {instance_type} {operating_system}")


def fetch_aws_gp3_price(pricing: Any) -> Price:
    filters = [
        {"Type": "TERM_MATCH", "Field": "regionCode", "Value": AWS_REGION},
        {"Type": "TERM_MATCH", "Field": "productFamily", "Value": "Storage"},
        {"Type": "TERM_MATCH", "Field": "volumeApiName", "Value": "gp3"},
    ]
    prices = [price for product in aws_products(pricing, filters) if (price := aws_ondemand_price(product, "GB-Mo"))]
    if not prices:
        raise PricingError("No AWS gp3 storage rate")
    return select_current_price(prices, "AWS gp3 storage")


def azure_pages(session: requests.Session, url: str, **kwargs: Any) -> Iterator[dict[str, Any]]:
    params = kwargs.pop("params", None)
    while url:
        payload = get_json(session, url, params=params, **kwargs)
        yield payload
        url = payload.get("nextLink") or payload.get("NextPageLink") or ""
        params = None


def azure_sku_restricted(item: dict[str, Any], region: str) -> bool:
    for restriction in item.get("restrictions", []):
        locations = set(restriction.get("values", []))
        locations.update(restriction.get("restrictionInfo", {}).get("locations", []))
        if restriction.get("type") == "Location" and (not locations or region in locations):
            return True
    return False


def parse_azure_sizes(items: Iterable[dict[str, Any]]) -> list[VmSize]:
    sizes: list[VmSize] = []
    for item in items:
        if item.get("resourceType") != "virtualMachines" or AZURE_REGION not in item.get("locations", []):
            continue
        if azure_sku_restricted(item, AZURE_REGION):
            continue
        capabilities = {cap["name"]: cap.get("value") for cap in item.get("capabilities", [])}
        if capabilities.get("CpuArchitectureType", "x64").lower() not in {"x64", "x86_64", "amd64"}:
            continue
        if "vCPUs" not in capabilities or "MemoryGB" not in capabilities:
            continue
        vcpu = int(capabilities["vCPUs"])
        memory = decimal(capabilities["MemoryGB"], f"Azure {item.get('name')} memory")
        if (vcpu, memory) in TARGET_SHAPES:
            sizes.append(VmSize(item["name"], vcpu, memory))
    return sorted({size.name: size for size in sizes}.values(), key=lambda item: (item.vcpu, item.name))


def discover_azure_sizes(session: requests.Session, subscription_id: str) -> list[VmSize]:
    credential = DefaultAzureCredential()
    token = credential.get_token("https://management.azure.com/.default").token
    url = f"https://management.azure.com/subscriptions/{subscription_id}/providers/Microsoft.Compute/skus"
    params = {"api-version": AZURE_SKUS_API_VERSION, "$filter": f"location eq '{AZURE_REGION}'"}
    headers = {"Authorization": f"Bearer {token}"}
    items: list[dict[str, Any]] = []
    for page in azure_pages(session, url, params=params, headers=headers):
        items.extend(page.get("value", []))
    return parse_azure_sizes(items)


def fetch_azure_retail_items(session: requests.Session, filter_query: str) -> list[dict[str, Any]]:
    params = {
        "api-version": "2023-01-01-preview",
        "currencyCode": "'AUD'",
        "$filter": filter_query,
        "meterRegion": "'primary'",
    }
    items: list[dict[str, Any]] = []
    for page in azure_pages(session, AZURE_RETAIL_URL, params=params):
        items.extend(page.get("Items", []))
    return items


def select_azure_compute_prices(items: Iterable[dict[str, Any]], operating_system: str) -> dict[str, Price]:
    grouped: dict[str, list[Price]] = {}
    for item in items:
        text = f"{item.get('productName', '')} {item.get('meterName', '')} {item.get('skuName', '')}".lower()
        product_name = str(item.get("productName", ""))
        is_windows = "windows" in product_name.lower()
        if (operating_system == OS_WINDOWS) != is_windows:
            continue
        if not product_name.startswith("Virtual Machines "):
            continue
        if any(excluded in text for excluded in ("spot", "low priority")):
            continue
        if item.get("type") != "Consumption" or item.get("unitOfMeasure") != "1 Hour":
            continue
        name = item.get("armSkuName")
        if name:
            grouped.setdefault(name, []).append(
                Price(
                    decimal(item.get("retailPrice"), f"Azure {name} price"),
                    str(item.get("effectiveStartDate") or "")[:10],
                    str(item.get("effectiveEndDate") or "")[:10],
                )
            )

    selected: dict[str, Price] = {}
    for name, prices in grouped.items():
        selected[name] = select_current_price(prices, f"Azure {operating_system} {name}")
    return selected


def select_azure_windows_prices(items: Iterable[dict[str, Any]]) -> dict[str, Price]:
    """Backward-compatible wrapper used by callers and tests."""
    return select_azure_compute_prices(items, OS_WINDOWS)


def select_azure_rhel_license_prices(items: Iterable[dict[str, Any]]) -> dict[int, Price]:
    grouped: dict[int, list[Price]] = {2: [], 4: []}
    for item in items:
        meter_name = str(item.get("meterName", ""))
        for vcpu in grouped:
            if (
                meter_name == f"{vcpu} vCPU VM License"
                and item.get("type") == "Consumption"
                and item.get("unitOfMeasure") == "1 Hour"
            ):
                price = decimal(item.get("retailPrice"), f"Azure RHEL {vcpu} vCPU licence")
                if price > 0:
                    grouped[vcpu].append(Price(
                        price,
                        str(item.get("effectiveStartDate") or "")[:10],
                        str(item.get("effectiveEndDate") or "")[:10],
                    ))
    return {
        vcpu: select_current_price(prices, f"Azure RHEL {vcpu} vCPU licence")
        for vcpu, prices in grouped.items()
    }


def select_azure_e10_price(items: Iterable[dict[str, Any]]) -> Price:
    prices: list[Price] = []
    for item in items:
        if (
            item.get("type") == "Consumption"
            and item.get("meterName") == "E10 LRS Disk"
            and item.get("unitOfMeasure") in {"1/Month", "1 Month"}
        ):
            prices.append(
                Price(
                    decimal(item.get("retailPrice"), "Azure E10 price"),
                    str(item.get("effectiveStartDate") or "")[:10],
                    str(item.get("effectiveEndDate") or "")[:10],
                )
            )
    if not prices:
        raise PricingError("No Azure E10 LRS disk rate")
    return select_current_price(prices, "Azure E10 LRS disk")


def build_row(
    provider: str,
    region: str,
    size: VmSize,
    compute_hourly: Decimal,
    disk_name: str,
    disk_monthly: Decimal,
    hours: Decimal,
    effective_date: str,
    fx_rate: str = "",
    fx_date: str = "",
    operating_system: str = OS_WINDOWS,
) -> ResultRow:
    total_monthly = compute_hourly * hours + disk_monthly
    return ResultRow(
        provider=provider,
        region=region,
        instance_type=size.name,
        vcpu=size.vcpu,
        memory_gib=size.memory_gib,
        compute_hourly_aud=compute_hourly,
        disk=disk_name,
        disk_monthly_aud=disk_monthly,
        total_hourly_aud=compute_hourly + disk_monthly / hours,
        total_monthly_aud=total_monthly,
        currency="AUD",
        pricing_effective_date=effective_date,
        fx_aud_usd=fx_rate,
        fx_date=fx_date,
        operating_system=operating_system,
        vm_shape=f"{size.vcpu} vCPU / {display_value(size.memory_gib)} GiB RAM",
    )


def collect_aws(
    profile: str | None, hours: Decimal, session: requests.Session, operating_systems: Sequence[str]
) -> list[ResultRow]:
    ec2, pricing = aws_clients(profile)
    sizes = discover_aws_sizes(ec2)
    if not sizes:
        raise PricingError("AWS returned no matching x86-64 instance types in Sydney")
    aud_usd, fx_date = fetch_rba_usd_rate(session)
    gp3 = fetch_aws_gp3_price(pricing)
    disk_monthly_aud = gp3.value * 128 / aud_usd
    rows: list[ResultRow] = []
    for operating_system in operating_systems:
        for size in sizes:
            compute = fetch_aws_compute_price(pricing, size.name, operating_system)
            if compute is None:
                continue
            rows.append(
                build_row(
                    "AWS", AWS_REGION, size, compute.value / aud_usd, "128 GiB gp3", disk_monthly_aud,
                    hours, compute.effective_date, str(aud_usd), fx_date, operating_system,
                )
            )
    if not rows:
        raise PricingError("AWS returned no matching instance types with the requested on-demand OS rate")
    return rows


def collect_azure(
    subscription_id: str, hours: Decimal, session: requests.Session, operating_systems: Sequence[str]
) -> list[ResultRow]:
    sizes = discover_azure_sizes(session, subscription_id)
    if not sizes:
        raise PricingError("Azure returned no matching x64 VM sizes in Australia East")
    vm_filter = f"serviceName eq 'Virtual Machines' and armRegionName eq '{AZURE_REGION}' and priceType eq 'Consumption'"
    retail_items = fetch_azure_retail_items(session, vm_filter)
    rhel_licences: dict[int, Price] = {}
    if OS_RHEL in operating_systems:
        rhel_filter = "productName eq 'Red Hat Enterprise Linux' and priceType eq 'Consumption'"
        rhel_licences = select_azure_rhel_license_prices(fetch_azure_retail_items(session, rhel_filter))
    disk_filter = (
        f"serviceName eq 'Storage' and armRegionName eq '{AZURE_REGION}' "
        "and productName eq 'Standard SSD Managed Disks' and priceType eq 'Consumption'"
    )
    disk = select_azure_e10_price(fetch_azure_retail_items(session, disk_filter))
    rows: list[ResultRow] = []
    for operating_system in operating_systems:
        base_os = OS_WINDOWS if operating_system == OS_WINDOWS else OS_LINUX
        prices = select_azure_compute_prices(retail_items, base_os)
        for size in sizes:
            if size.name not in prices:
                continue
            compute_price = prices[size.name]
            if operating_system == OS_RHEL:
                licence = rhel_licences[size.vcpu]
                compute_price = Price(
                    compute_price.value + licence.value,
                    max(compute_price.effective_date, licence.effective_date),
                )
            rows.append(build_row(
                "Azure", AZURE_REGION, size, compute_price.value, "128 GiB Standard SSD LRS (E10)",
                disk.value, hours, compute_price.effective_date, operating_system=operating_system,
            ))
    if not rows:
        raise PricingError("Azure returned no matching VM sizes with the requested PAYG OS rate")
    return rows


EXCEL_COLUMNS = [
    ("Provider", "provider"),
    ("Operating System", "operating_system"),
    ("VM Shape", "vm_shape"),
    ("Region", "region"),
    ("Instance / VM Size", "instance_type"),
    ("vCPU", "vcpu"),
    ("RAM (GiB)", "memory_gib"),
    ("Compute / Hour (AUD)", "compute_hourly_aud"),
    ("Disk", "disk"),
    ("Disk / Month (AUD)", "disk_monthly_aud"),
    ("Total / Hour (AUD)", "total_hourly_aud"),
    ("Total / Month (AUD)", "total_monthly_aud"),
    ("Currency", "currency"),
    ("Price Effective Date", "pricing_effective_date"),
    ("AUD/USD FX Rate", "fx_aud_usd"),
    ("FX Date", "fx_date"),
    ("Status", "status"),
]


def display_value(value: Any) -> str:
    if isinstance(value, Decimal):
        return f"{value:.6f}".rstrip("0").rstrip(".")
    return str(value)


def excel_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    return value


def write_excel(rows: Sequence[ResultRow], path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    shape_fills = {
        "2 vCPU / 8 GiB RAM": PatternFill("solid", fgColor="EAF3F8"),
        "4 vCPU / 16 GiB RAM": PatternFill("solid", fgColor="FFF2CC"),
    }
    group_border = Border(top=Side(style="medium", color="4472C4"))
    currency_fields = {
        "compute_hourly_aud", "disk_monthly_aud", "total_hourly_aud", "total_monthly_aud"
    }

    for provider in ("AWS", "Azure"):
        sheet = workbook.create_sheet(provider)
        provider_rows = [row for row in rows if row.provider == provider]
        headers = [heading for heading, _ in EXCEL_COLUMNS]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in provider_rows:
            values = asdict(row)
            sheet.append([excel_value(values[field]) for _, field in EXCEL_COLUMNS])

        previous_group: tuple[str, str] | None = None
        for row_index, row in enumerate(provider_rows, start=2):
            group = (row.operating_system, row.vm_shape)
            fill = shape_fills.get(row.vm_shape)
            for cell in sheet[row_index]:
                if fill:
                    cell.fill = fill
                if group != previous_group:
                    cell.border = group_border
                    cell.font = Font(bold=True)
            previous_group = group

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        sheet.row_dimensions[1].height = 24
        for column_index, (heading, field) in enumerate(EXCEL_COLUMNS, start=1):
            letter = sheet.cell(1, column_index).column_letter
            observed = [heading, *(str(sheet.cell(row, column_index).value or "") for row in range(2, sheet.max_row + 1))]
            sheet.column_dimensions[letter].width = min(max(len(value) for value in observed) + 2, 42)
            if field in currency_fields:
                for row_index in range(2, sheet.max_row + 1):
                    sheet.cell(row_index, column_index).number_format = 'A$0.000000'
            elif field == "memory_gib":
                for row_index in range(2, sheet.max_row + 1):
                    sheet.cell(row_index, column_index).number_format = '0.##'
            elif field == "fx_aud_usd":
                for row_index in range(2, sheet.max_row + 1):
                    sheet.cell(row_index, column_index).number_format = '0.000000'

        if provider_rows:
            table = Table(displayName=f"{provider}Pricing", ref=sheet.dimensions)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
                showRowStripes=True, showColumnStripes=False,
            )
            sheet.add_table(table)
    workbook.save(path)


def print_table(rows: Sequence[ResultRow]) -> None:
    headers = ["Instance", "Compute/hr AUD", "Disk/mo AUD", "Total/mo AUD"]
    for provider in ("AWS", "Azure"):
        provider_rows = [row for row in rows if row.provider == provider]
        if not provider_rows:
            continue
        print(f"\n{provider}")
        for operating_system in ALL_OPERATING_SYSTEMS:
            for vcpu, memory in sorted(TARGET_SHAPES):
                group_rows = [
                    row for row in provider_rows
                    if row.operating_system == operating_system
                    and row.vcpu == vcpu
                    and row.memory_gib == memory
                ]
                if not group_rows:
                    continue
                print(f"\n  {operating_system} — {vcpu} vCPU / {display_value(memory)} GiB RAM")
                data = [
                    [row.instance_type, f"{row.compute_hourly_aud:.4f}",
                     f"{row.disk_monthly_aud:.2f}", f"{row.total_monthly_aud:.2f}"]
                    for row in group_rows
                ]
                widths = [
                    max(len(str(value)) for value in [header, *(line[index] for line in data)])
                    for index, header in enumerate(headers)
                ]
                print("  " + "  ".join(header.ljust(widths[index]) for index, header in enumerate(headers)))
                print("  " + "  ".join("-" * width for width in widths))
                for line in data:
                    print("  " + "  ".join(str(value).ljust(widths[index]) for index, value in enumerate(line)))


def cheapest_per_provider(rows: Sequence[ResultRow], limit: int) -> list[ResultRow]:
    selected: list[ResultRow] = []
    for provider in ("AWS", "Azure"):
        operating_systems = [os_name for os_name in ALL_OPERATING_SYSTEMS if any(
            row.provider == provider and row.operating_system == os_name for row in rows
        )]
        for operating_system in operating_systems:
            for vcpu, memory_gib in sorted(TARGET_SHAPES):
                shape_rows = sorted(
                    (
                        row for row in rows
                        if row.provider == provider
                        and row.operating_system == operating_system
                        and row.vcpu == vcpu
                        and row.memory_gib == memory_gib
                    ),
                    key=lambda row: (row.total_monthly_aud, row.instance_type),
                )
                selected.extend(shape_rows[:limit])
    return selected


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--aws-profile", help="AWS shared-config profile (default: normal credential chain)")
    parser.add_argument("--azure-subscription-id", default=os.getenv("AZURE_SUBSCRIPTION_ID"))
    parser.add_argument("--output", type=Path, default=Path("vm_pricing_aud.xlsx"))
    parser.add_argument("--hours-per-month", type=Decimal, default=DEFAULT_HOURS)
    parser.add_argument(
        "--top", type=int, default=5,
        help="Cheapest rows per provider, OS, and VM shape (default: 5)",
    )
    parser.add_argument(
        "--os", choices=("windows", "linux", "rhel", "all"), default="all",
        help="Operating-system pricing to include (default: all three)",
    )
    parser.add_argument("--allow-partial", action="store_true", help="Write results if one provider fails")
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    if not args.azure_subscription_id:
        print("error: supply --azure-subscription-id or AZURE_SUBSCRIPTION_ID", file=sys.stderr)
        return 2
    if args.hours_per_month <= 0:
        print("error: --hours-per-month must be positive", file=sys.stderr)
        return 2
    if args.top <= 0:
        print("error: --top must be positive", file=sys.stderr)
        return 2
    if args.output.suffix.lower() != ".xlsx":
        print("error: --output must use the .xlsx extension", file=sys.stderr)
        return 2

    session = http_session()
    os_choices = {"windows": OS_WINDOWS, "linux": OS_LINUX, "rhel": OS_RHEL}
    operating_systems = ALL_OPERATING_SYSTEMS if args.os == "all" else (os_choices[args.os],)
    rows: list[ResultRow] = []
    failures: list[str] = []
    collectors = (
        ("AWS", lambda: collect_aws(args.aws_profile, args.hours_per_month, session, operating_systems)),
        ("Azure", lambda: collect_azure(args.azure_subscription_id, args.hours_per_month, session, operating_systems)),
    )
    for provider, collect in collectors:
        try:
            rows.extend(collect())
        except Exception as exc:  # SDK exceptions are intentionally surfaced as provider failures.
            failures.append(f"{provider}: {exc}")
            if not args.allow_partial:
                break

    if failures and (not args.allow_partial or not rows):
        print("Pricing collection failed:\n  " + "\n  ".join(failures), file=sys.stderr)
        return 1
    rows.sort(key=lambda row: (row.total_monthly_aud, row.provider, row.instance_type))
    rows = cheapest_per_provider(rows, args.top)
    write_excel(rows, args.output)
    print_table(rows)
    print(f"\nWrote {len(rows)} rows to {args.output}")
    print(f"Retrieved at {datetime.now(timezone.utc).isoformat(timespec='seconds')}")
    if failures:
        print("Partial result warnings:\n  " + "\n  ".join(failures), file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
